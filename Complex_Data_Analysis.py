import openpyxl

wb = openpyxl.load_workbook("abc.xlsx")
del wb['Notes']
s1 = wb['Service 1']
s2 = wb['Service 2']
s3 = wb['Service 3']
s4 = wb['Master view']
row1 = s1.max_row + 1
row2 = s2.max_row + 1
master_row2 = row1 + row2 -1
row3 = s3.max_row + 1
master_row3 = row3 + master_row2 -1
col1 = s1.max_column + 1
col2 = s2.max_column + 1
col3 = s3.max_column + 1
sh3_altid = 0

for i in range(1, row1):
    b_id = s1.cell(i, 1).value
    b_date = str(s1.cell(i, 6).value)
    b_service = s1.cell(i, 3).value
    b_servicename = s1.cell(i, 9).value
    b_collected = s1.cell(i, 7).value
    #    date_time_obj = datetime.strptime(b_date, '%m/%d/%y %H:%M')
    if i!= 1:
        s1.cell(row=i, column=1, value=b_date)
        s1.cell(row=i, column=2, value=b_id)
        s1.cell(row=i, column=3, value=b_service)

    for j in range(4, col1):
        s1.cell(row=i, column=j, value=" ")
for i in range(row1,48227):
    s1.cell(row=i, column=1, value="sdfg")

for i in range(1, row2):
    b_date = str(s2.cell(i, 7).value)
    b_id = s2.cell(i, 1).value
    b_service = s2.cell(i, 2).value
    b_net = s2.cell(i, 3).value
    b_invoice = s2.cell(i, 4).value
    b_collected = s2.cell(i, 6).value
    i = i - 1
    if i != 0:
        s2.cell(row=i+1, column=1, value=b_date)
        s2.cell(row=i+1, column=2, value=b_id)
        s2.cell(row=i+1, column=3, value=b_service)
        s4.cell(row=i + row1, column=1, value=b_id)  # ID
        s4.cell(row=i + row1, column=2, value=b_service)  # SERVICE
        s4.cell(row=i+row1, column=4, value="NONE")  # sERVICE NAME
        s4.cell(row=i + row1, column=5, value=b_net)  # NET AMOUNT
        s4.cell(row=i + row1, column=6, value=b_invoice)  # iNVOICE STATUS
        s4.cell(row=i + row1, column=7, value=b_date)  # bILLED DATE
        s4.cell(row=i + row1, column=8, value=b_collected)  # COLLECETED DATE
    i = i + 1

    for j in range(4, col2):
        s2.cell(row=i, column=j, value=" ")

for i in range(1, row3):
    b_id = s3.cell(i, 1).value
    b_date = str(s3.cell(i, 9).value)
    b_service = s3.cell(i, 2).value
    s3.cell(row=i, column=1, value=b_date)
    s3.cell(row=i, column=3, value=b_service)
    if type(b_id) == str and i != 1:
        s3.cell(row=i, column=2, value=sh3_altid)
        sh3_altid = sh3_altid + 1
    elif type(b_id) == int and i != 1:
        s3.cell(row=i, column=2, value=b_id)
        sh3_altid = int(b_id) + 1
    else:
        s3.cell(row=i, column=2, value=b_id)

    for j in range(4, col3):
        s3.cell(row=i, column=j, value=" ")

for i in range(1,master_row3):
    if i<=row1:
        b_id = s1.cell(i, 1).value
        b_date = str(s1.cell(i, 6).value)
        b_service = s1.cell(i, 3).value
        b_servicename = s1.cell(i, 9).value
        b_collected = s1.cell(i, 7).value
        if i!=1:
            s4.cell(row=i, column=1, value=b_id)  # ID
            s4.cell(row=i, column=2, value=b_service)
            s4.cell(row=i, column=4, value=b_servicename)
            s4.cell(row=i, column=5, value=0)
            s4.cell(row=i, column=6, value="settled")
            s4.cell(row=i, column=7, value=b_date)
            s4.cell(row=i, column=8, value=b_collected)
    elif i<=master_row2 and i>row1:


wb.save("SidekickEDGE - Python Exercise.xlsx")
