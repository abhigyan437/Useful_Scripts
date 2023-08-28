def Reverse(lst):
    lst.reverse()
    return lst

import openpyxl as op
wb = op.load_workbook("Sample_Longlist_Data.xlsx")
content = wb['Sheet1']
col = 18
row = content.max_row
tuple_empty = []
count1 = 2
count2 = 1
while(count1<=row):
    content.cell(count1,4).value = str(content.cell(count1,3).value)+"-link"
    content.cell(count1,5).value = str(content.cell(count1,3).value)+"-storefrontlink"
    count1 = count1+1


count1 = 2
while(count1<=row):
    while(count2<=col):
        if not content.cell(count1, count2).value and content.cell(count1, count2).value!= 0:
            tuple_empty.append(count1)
            break
        if count2==col:
            count2 = 1
            break
        count2 = count2 +1
    count1 = count1+1
tuple_empt = Reverse(tuple_empty)

num = len(tuple_empty)

row = row-num

for qwe in tuple_empt:
    content.delete_rows(idx=qwe,amount=1)
count1 = 2
tuple_negative = []

while(count1<=row):
    if content.cell(count1,13).value>10 :
        tuple_negative.append(count1)
    count1 = count1+1
num = len(tuple_negative)
row = row-num
tuple_negativ=Reverse(tuple_negative)
for qwe in tuple_negativ:
    content.delete_rows(idx=qwe, amount=1)
count1 = 2
tuple_rating1 = []
while(count1<=row):
    if content.cell(count1,15).value<10000 :
        tuple_rating1.append(count1)
    count1 = count1+1
num = len(tuple_rating1)
tuple_ating1 = Reverse(tuple_rating1)
row = row-num
for qwe in tuple_ating1:
    content.delete_rows(idx=qwe, amount=1)

count1 = 2
tuple_rating2 = []
while(count1<=row):
    if content.cell(count1,16).value<10000 :
        tuple_rating2.append(count1)
    count1 = count1+1
tuple_ating2 = Reverse(tuple_rating2)
num = len(tuple_rating2)
row = row-num
for qwe in tuple_ating2:
    content.delete_rows(idx=qwe, amount=1)

wb.save("result.xlsx")
