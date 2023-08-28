import openpyxl

wb = openpyxl.load_workbook('abc.xlsx')

# Sheet 1 modification
s1 = wb['Service 1']
row1 = s1.max_row + 1
col1 = s1.max_column + 1
for i in range(1, row1):
    b_date = s1.cell(i, 6).value
    b_id = s1.cell(i, 1).value
    b_service = s1.cell(i, 3).value
    s1.cell(row=i, column=1, value=b_date)
    s1.cell(row=i, column=2, value=b_id)
    s1.cell(row=i, column=3, value=b_service)
    for j in range(4, col1):
        s1.cell(row=i, column=j, value=" ")

# sheet2 modifcation
s2 = wb['Service 2']
row2 = s2.max_row + 1
col2 = s2.max_column + 1
for i in range(2, row2):
    b_date = s2.cell(i, 5).value
    b_id = s2.cell(i, 1).value
    b_service = s2.cell(i, 2).value
    s1.cell(row=i, column=1, value=b_date)
    s1.cell(row=i, column=2, value=b_id)
    s1.cell(row=i, column=3, value=b_service)
    for j in range(4, col2):
        s1.cell(row=i, column=j, value=" ")

wb.save("result.xlsx")
